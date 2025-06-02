import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
import sqlite3
import re
import os

def get_brand_map():
    return {
        "72": "Owen Barry",
        "52": "Crockett&Jones",
        "69": "PYRENEX",
        "01": "BARBARIAN",
        "71": "BEORMA",
        "04": "SETTLER",
        "10": "McROSTIE",
        "70": "William Lockie",
        "53": "PASHMINA",
        "30": "WHITEHOUSE COX",
        "08": "Anderson & Co",
        "68": "FILSON",
        "59": "CANADA GOOSE",
        "35": "Northern Watters Knitwear"
    }

def process_file(filename):
    match = re.search(r"(\d{4})年(\d{1,2})月", filename)
    if not match:
        messagebox.showerror("エラー", "ファイル名に年と月が含まれていません")
        return
    year, month = int(match.group(1)), int(match.group(2))

    db_name = "frame.db"
    conn = sqlite3.connect(db_name)
    cursor = conn.cursor()
    cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='orders'")
    if not cursor.fetchone():
        empty_df = pd.DataFrame(columns=[
            "受注№", "受注日", "年齢", "性別", "品番", "品名", "色", "サイズ",
            "税抜金額", "税込金額", "居住地", "支払方法", "年", "月", "点数", "ギフト", "ブランド"
        ])
        empty_df.to_sql("orders", conn, if_exists="fail", index=False)

    df_existing = pd.read_sql("SELECT * FROM orders", conn)
    df = pd.read_excel(filename, sheet_name=0)

    # 列の順番に基づいてデータを選択
    column_indices = {
        "受注日": 0,
        "受注№": 1,
        "年齢": 4,
        "性別": 5,
        "品番": 6,
        "品名": 7,
        "色": 8,
        "サイズ": 9,
        "税抜金額": 10,
        "税込金額": 11,
        "居住地": 12,
        "支払方法": 13,
        "ギフト": 14
    }

    # 必要な列をインデックスで選択
    selected_columns = [index for index in column_indices.values() if index < len(df.columns)]
    df_selected = df.iloc[:, selected_columns]
    print("デバッグ情報: 選択された列")
    print(df_selected.columns)

    # 列名を設定
    df_selected.columns = [name for name, index in column_indices.items() if index < len(df.columns)]

    # デバッグ情報: 選択された列を確認
    #print("デバッグ情報: 選択された列")
    #print(df_selected.columns)

    # 必要に応じてデータ型を変換
    df_selected.loc[df_selected["品番"].notna(), "品番"] = (
        df_selected.loc[df_selected["品番"].notna(), "品番"]
        .astype(str).str.zfill(12)
    )
    df_selected = df_selected.dropna(subset=["受注№", "受注日"])
    df_selected["受注日"] = pd.to_numeric(df_selected["受注日"], errors="coerce").astype("Int64")
    df_selected["年齢"] = pd.to_numeric(df_selected["年齢"], errors="coerce").astype("Int64")
    df_selected["税込金額"] = pd.to_numeric(df_selected["税込金額"], errors="coerce")
    df_selected["税抜金額"] = df_selected["税抜金額"].round(0).astype("Int64")

    # "品名" が "送料" の場合に "品番" を "979900010199" に設定
    if "品名" in df_selected.columns and "品番" in df_selected.columns:
        df_selected.loc[df_selected["品名"] == "送料", "品番"] = "979900010199"

    # ブランド名追加
    brand_map = get_brand_map()
    df_selected["ブランド"] = df_selected["品番"].astype(str).str[:2].map(brand_map).fillna("その他")


    # マイナス金額補完処理（前の販売日コピー）
    last_positive_date = None
    for idx, row in df_selected.iterrows():
        if row["税込金額"] >= 0:
            last_positive_date = row["受注日"]
        else:
            if last_positive_date is not None:
                df_selected.at[idx, "受注日"] = last_positive_date

    # その他の処理を続行
    df_selected["年"] = year
    df_selected["月"] = month
    df_selected["点数"] = df_selected["税込金額"].apply(lambda x: -1 if x < 0 else 1)

    # ギフト対象品番
    gift_items = ["909900220199", "909900250199", "699900030199"]
    # 品番によるギフト判定
    gift_orders_by_item = df_selected[df_selected["品番"].astype(str).isin(gift_items)]["受注№"].unique()
    # ギフトリボンに「リボン」が含まれるかどうか（NaNを除外して文字列として検索）
    has_ribbon = df_selected["ギフト"].fillna("").astype(str).str.contains("リボン")
    # ギフトフラグ設定
    df_selected["ギフト"] = (
        df_selected["受注№"].isin(gift_orders_by_item) | has_ribbon
    ).astype(int)


    compare_cols = ["受注№", "品番", "税込金額"]
    compare_cols = [col for col in compare_cols if col in df_selected.columns and col in df_existing.columns]
    df_selected[compare_cols] = df_selected[compare_cols].astype(str)
    df_existing[compare_cols] = df_existing[compare_cols].astype(str)

    df_selected = pd.merge(
        df_selected,
        df_existing[compare_cols].drop_duplicates(),
        on=compare_cols,
        how="left",
        indicator=True
    )
    df_selected = df_selected[df_selected["_merge"] == "left_only"].drop(columns=["_merge"])

    if df_selected.empty:
        messagebox.showinfo("結果", "✅ 新規データはありませんでした。")
    else:
        df_selected.to_sql("orders", conn, if_exists="append", index=False)
        diff_csv = f"{year}年{month}月_差分出力.csv"
        #df_selected.to_csv(diff_csv, index=False, encoding="utf-8-sig")
        messagebox.showinfo("完了", f"✅ 新規データ {len(df_selected)} 件を登録しました。差分CSV: {diff_csv}")

    df_all = pd.read_sql("SELECT * FROM orders", conn)
    df_all.to_csv("frame_全データ出力.csv", index=False, encoding="utf-8-sig")
    conn.close()

def process_fukuoka_file(filename):
    import re
    import pandas as pd
    import sqlite3
    import os
    from tkinter import messagebox

    # ファイル名から年・月を抽出（例: データ出力202404 → 2024, 4）
    match = re.search(r"データ出力(\d{4})(\d{2})", os.path.basename(filename))
    if not match:
        messagebox.showerror("エラー", "ファイル名に 'データ出力yyyymm' の形式がありません")
        return
    year, month = int(match.group(1)), int(match.group(2))

    db_name = "framefukuoka.db"
    conn = sqlite3.connect(db_name)
    cursor = conn.cursor()
    # テーブル作成（なければ）
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS orders (
            品番 TEXT,
            品名 TEXT,
            色 TEXT,
            サイズ TEXT,
            税抜金額 INTEGER,
            税込金額 INTEGER,
            点数 INTEGER,
            年 INTEGER,
            月 INTEGER,
            ブランド TEXT
        )
    """)
    conn.commit()

    # Excel読込
    try:
        df = pd.read_excel(filename, dtype=str)
        # 追加: 社員名が存在する場合「ｗｅｂ」以外のみ残す
        if "社員名" in df.columns:
            df = df[df["社員名"] != "ｗｅｂ"]
    except Exception as e:
        messagebox.showerror("エラー", f"Excelファイルの読込に失敗しました: {e}")
        conn.close()
        return

    # 列名マッピング
    col_map = {
        "品番": "商品CD",
        "品名": "商品名",
        "色": "カラー",
        "サイズ": "サイズ",
        "税抜金額": "プロパー金額",
        "点数": "売上数"
    }
    # 必要な列のみ抽出
    missing_cols = [v for v in col_map.values() if v not in df.columns]
    if missing_cols:
        messagebox.showerror("エラー", f"列名が不足しています: {missing_cols}")
        conn.close()
        return
    df_selected = df[[col_map[k] for k in col_map]].copy()
    df_selected.columns = list(col_map.keys())

    # データ型変換
    df_selected["税抜金額"] = pd.to_numeric(df_selected["税抜金額"], errors="coerce").fillna(0).astype(int)
    df_selected["点数"] = pd.to_numeric(df_selected["点数"], errors="coerce").fillna(0).astype(int)

    # 税込金額計算
    def calc_tax(price, y, m):
        if y < 2014 or (y == 2014 and m <= 3):
            rate = 1.05
        elif (y == 2014 and m >= 4) or (2014 < y < 2019) or (y == 2019 and m <= 9):
            rate = 1.08
        else:
            rate = 1.10
        return int(round(price * rate))
    df_selected["税込金額"] = df_selected["税抜金額"].apply(lambda x: calc_tax(x, year, month))

    # 年・月を追加
    df_selected["年"] = year
    df_selected["月"] = month

    # ブランド名追加（関数化したget_brand_mapを利用）
    brand_map = get_brand_map()
    df_selected["ブランド"] = df_selected["品番"].astype(str).str[:2].map(brand_map).fillna("その他")

    # DBへINSERT
    try:
        df_selected.to_sql("orders", conn, if_exists="append", index=False)
        messagebox.showinfo("完了", f"✅ {len(df_selected)} 件を登録しました。")
    except Exception as e:
        messagebox.showerror("エラー", f"DB登録に失敗しました: {e}")

    df_all = pd.read_sql("SELECT * FROM orders", conn)
    df_all.to_csv("framefukuoka_全データ出力.csv", index=False, encoding="utf-8-sig")
    conn.close()

def upload_file():
    file_path = filedialog.askopenfilename(
        title="Excelファイルを選択してください",
        filetypes=[("Excel Files", "*.xlsx *.xls")]
    )
    if file_path:
        process_file(file_path)

def upload_files():
    file_paths = filedialog.askopenfilenames(
        title="Excelファイルを選択してください",
        filetypes=[("Excel Files", "*.xlsx *.xls")]
    )
    if file_paths:
        for file_path in file_paths:
            process_file(file_path)
        messagebox.showinfo("完了", f"✅ {len(file_paths)} 件のファイルを処理しました。")

def upload_fukuoka_file():
    file_path = filedialog.askopenfilename(
        title="福岡用Excelファイルを選択してください",
        filetypes=[("Excel Files", "*.xlsx *.xls")]
    )
    if file_path:
        process_fukuoka_file(file_path)

def upload_fukuoka_files():
    file_paths = filedialog.askopenfilenames(
        title="福岡用Excelファイルを選択してください",
        filetypes=[("Excel Files", "*.xlsx *.xls")]
    )
    if file_paths:
        for file_path in file_paths:
            process_fukuoka_file(file_path)
        messagebox.showinfo("完了", f"✅ {len(file_paths)} 件のファイルを処理しました。")

def export_data(start_product_code, end_product_code):
    start_year = int(start_year_var.get())
    start_month = int(start_month_var.get())
    end_year = int(end_year_var.get())
    end_month = int(end_month_var.get())
    start_key = start_year * 100 + start_month
    end_key = end_year * 100 + end_month

    # ダウンロードフォルダのパスを取得
    folder = os.path.join(os.path.expanduser("~"), "Downloads")

    conn = sqlite3.connect("frame.db")
    df = pd.read_sql("SELECT * FROM orders", conn)

    # デバッグ情報: データ型を確認
    print("デバッグ情報: データ型")
    print(df.dtypes)

    # 年と月を数値型に変換
    df["年"] = pd.to_numeric(df["年"], errors="coerce")
    df["月"] = pd.to_numeric(df["月"], errors="coerce")

    # 欠損値を確認
    print("デバッグ情報: 欠損値の数")
    print(df[["年", "月"]].isna().sum())

    # 欠損値を削除
    df = df.dropna(subset=["年", "月"])

    # 年月キーの計算
    df["年月キー"] = (df["年"].astype(int) * 100 + df["月"].astype(int)).astype("Int64")

    # デバッグ情報: 年月キーの計算結果を確認
    print("デバッグ情報: 年月キーの計算結果")
    print(df[["年", "月", "年月キー"]].head())

    # 年月でフィルタリング
    filtered_df = df[(df["年月キー"] >= start_key) & (df["年月キー"] <= end_key)].drop(columns=["年月キー"])

    # "品番" で範囲フィルタリング（入力がある場合のみ）
    if start_product_code and end_product_code:
        filtered_df = filtered_df[
            (filtered_df["品番"] >= start_product_code) & (filtered_df["品番"] <= end_product_code)
        ]

    conn.close()
    if filtered_df.empty:
        messagebox.showinfo("結果", "該当期間または指定された品番のデータはありませんでした。")
        return

    # 列の並びを変更
    desired_order = [
        "受注№", "年", "月", "受注日", "年齢", "性別", "品番", "品名", "色", "サイズ",
        "税抜金額", "税込金額", "点数", "ギフト", "支払方法","居住地" ,"ブランド"   ]
    filtered_df = filtered_df[desired_order]

    # ファイル名と保存先パスを設定
    filename = f"{start_year}年{start_month}月-{end_year}年{end_month}月_データ.xlsx"
    full_path = os.path.join(folder, filename)

    # ✅ 数値型に変換（ここを追加）
    num_cols = ["受注日", "年齢", "税抜金額", "税込金額", "点数"]
    for col in num_cols:
        if col in filtered_df.columns:
            filtered_df[col] = pd.to_numeric(filtered_df[col], errors="coerce")


    # データをExcelに保存
    filtered_df.to_excel(full_path, index=False)
    messagebox.showinfo("完了", f"✅ データをExcelに出力しました：\n{full_path}")

def get_leather_type(product_name):
    # 革の種類のキーワードリスト
    leather_keywords = [
        "SAMPL", "OXFORD", "LON/BRI", "BRITISH COUNTRY", "SHRUNKEN", "SAFARI", "DERBY", "REGENT",
        "LON", "CORDOVAN", "ST.JAMES", "GASTON", "SHRUNKEN", "HAMPSTEAD",
        "VIN BR", "HORWEEN/BRI", "BR/RUS", "PLAITED", "VTC BADALASSI",
        "BADALASSI", "NATUR/BR", "NATUR", "CAN/TUS", "TUS", "RUSSET", "PLAI",
        "FIX LEATHER", "WEBBNG", "WEBBING", "PASTURE SUEDE", "STIRRUP", "LEATHER BALM", "BR2", "BRI"
    ]
    for keyword in leather_keywords:
        # 完全一致のみ判定（部分一致しない）
        if str(product_name) == keyword:
            return keyword
    # 部分一致で判定（LONはLONDONなど先頭一致、LONGは除外）
    for keyword in leather_keywords:
        if keyword == "LON":
            s = str(product_name).upper()
            # LONDONなど先頭一致、LONGは除外
            if s.startswith("LON") and not s.startswith("LONG"):
                return keyword
            # ' LONDON'や' LON 'などもOK
            if re.search(r'\bLON', s) and not re.search(r'\bLONG', s):
                return keyword
        else:
            if keyword in str(product_name):
                return keyword
    return "その他"

def upload_product_excel():
    file_path = filedialog.askopenfilename(
        title="商品マスタExcelファイルを選択してください",
        filetypes=[("Excel Files", "*.xlsx *.xls")]
    )
    if not file_path:
        return
    # DB作成・接続
    db_name = "product.db"
    conn = sqlite3.connect(db_name)
    cursor = conn.cursor()
    # productsテーブル作成（なければ）
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS products (
            商品名 TEXT,
            品番CD TEXT,
            カラーNO TEXT,
            カラー名 TEXT,
            サイズ数計 INTEGER,
            金額 REAL,
            革の種類 TEXT,
            ブランド TEXT,
            PRIMARY KEY (品番CD, カラーNO)
        )
    """)
    conn.commit()
    # Excel読込
    try:
        df = pd.read_excel(file_path, dtype=str)
    except Exception as e:
        messagebox.showerror("エラー", f"Excelファイルの読込に失敗しました: {e}")
        conn.close()
        return
    # 必要な列のみ抽出
    required_cols = ["商品名", "品番CD", "カラーNO", "カラー名","サイズ数計"]
    if not all(col in df.columns for col in required_cols):
        messagebox.showerror("エラー", f"列名が不足しています: {required_cols}")
        conn.close()
        return
    df = df[required_cols].copy()
    # 金額・革の種類列追加
    # df["金額"] = None  ← この行は削除またはコメントアウト
    df["革の種類"] = df["商品名"].apply(get_leather_type)

    # ブランド名追加
    brand_map = get_brand_map()
    df["ブランド"] = df["品番CD"].astype(str).str[:2].map(brand_map).fillna("その他")

    # DB登録（UPSERT: 金額は常に更新しない）
    inserted_or_updated = 0
    for _, row in df.iterrows():
        try:
            cursor.execute("""
                INSERT INTO products (商品名, 品番CD, カラーNO, カラー名, サイズ数計, 革の種類,ブランド)
                VALUES (?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(品番CD, カラーNO) DO UPDATE SET
                    商品名=excluded.商品名,
                    カラー名=excluded.カラー名,
                    サイズ数計=excluded.サイズ数計,
                    革の種類=excluded.革の種類,
                    ブランド=excluded.ブランド
            """, (
                row["商品名"], row["品番CD"], row["カラーNO"], row["カラー名"], row["サイズ数計"],
                row["革の種類"], row["ブランド"]
            ))
            inserted_or_updated += 1
        except Exception as e:
            print(f"登録エラー: {e}")
    conn.commit()
    messagebox.showinfo("完了", f"✅ {inserted_or_updated} 件を登録（新規または更新）しました。")
    df_all = pd.read_sql("SELECT * FROM products", conn)
    df_all.to_csv("商品マスタ.csv", index=False, encoding="utf-8-sig")
    conn.close()

def upload_price_csv():
    file_path = filedialog.askopenfilename(
        title="商品マスタ金額CSVファイルを選択してください",
        filetypes=[("CSV Files", "*.csv")]
    )
    if not file_path:
        return
    db_name = "product.db"
    conn = sqlite3.connect(db_name)
    try:
        df_csv = pd.read_csv(file_path, dtype=str, encoding="cp932")
    except Exception as e:
        messagebox.showerror("エラー", f"CSVファイルの読込に失敗しました: {e}")
        conn.close()
        return
    # 必要な列名を確認
    if not ("商品コード" in df_csv.columns and "通常価格" in df_csv.columns):
        messagebox.showerror("エラー", "CSVに必要な列（商品コード、通常価格）がありません。")
        conn.close()
        return
    # DBから既存データ取得
    df_db = pd.read_sql("SELECT * FROM products", conn)
    update_count = 0
    for idx, row in df_csv.iterrows():
        code8 = str(row["商品コード"])[:8]
        price_str = row["通常価格"]
        try:
            price = float(price_str)
            price_wo_tax = round(price / 1.1)
        except (ValueError, TypeError):
            continue  # 金額が変換できない場合はスキップ
        # DB内で品番CDの上8桁が一致する行を抽出
        mask = df_db["品番CD"].str[:8] == code8
        if mask.any():
            df_db.loc[mask, "金額"] = price_wo_tax
            update_count += mask.sum()
    # DBへ反映
    try:
        # 一時テーブルに書き出し
        df_db.to_sql("products_tmp", conn, if_exists="replace", index=False)
        cursor = conn.cursor()
        # 本来のproductsテーブルを削除
        cursor.execute("DROP TABLE IF EXISTS products")
        # 主キー付きで再作成
        cursor.execute("""
            CREATE TABLE products (
                商品名 TEXT,
                品番CD TEXT,
                カラーNO TEXT,
                カラー名 TEXT,
                サイズ数計 INTEGER,
                金額 REAL,
                革の種類 TEXT,
                ブランド TEXT,
                PRIMARY KEY (品番CD, カラーNO)
            )
        """)
        # データを移し替え
        cursor.execute("""
            INSERT INTO products (商品名, 品番CD, カラーNO, カラー名, サイズ数計, 金額, 革の種類, ブランド)
            SELECT 商品名, 品番CD, カラーNO, カラー名, サイズ数計, 金額, 革の種類, ブランド FROM products_tmp
        """)
        cursor.execute("DROP TABLE IF EXISTS products_tmp")
        conn.commit()
        messagebox.showinfo("完了", f"✅ {update_count} 件の金額を更新しました。")
    except Exception as e:
        messagebox.showerror("エラー", f"DB更新に失敗しました: {e}")
    df_all = pd.read_sql("SELECT * FROM products", conn)
    df_all.to_csv("商品マスタ.csv", index=False, encoding="utf-8-sig")
    conn.close()

def download_sales_summary():
    import pandas as pd
    import sqlite3
    import os

    # 年月・品番範囲の取得
    start_year = int(start_year_var.get())
    start_month = int(start_month_var.get())
    end_year = int(end_year_var.get())
    end_month = int(end_month_var.get())
    start_key = start_year * 100 + start_month
    end_key = end_year * 100 + end_month
    start_product_code = start_product_code_var.get()
    end_product_code = end_product_code_var.get()

    db_mode = db_select_var.get()  # ラジオボタンの値を取得

    # DB接続
    conn_prod = sqlite3.connect("product.db")
    try:
        df_products = pd.read_sql("SELECT * FROM products", conn_prod)
    except Exception as e:
        messagebox.showerror("エラー", f"商品マスタDB読込エラー: {e}")
        conn_prod.close()
        return

    # --- データ取得 ---
    if db_mode == "WEB":
        conn_frame = sqlite3.connect("frame.db")
        try:
            df_orders = pd.read_sql("SELECT * FROM orders", conn_frame)
        except Exception as e:
            messagebox.showerror("エラー", f"WEB用DB読込エラー: {e}")
            conn_prod.close()
            conn_frame.close()
            return
        conn_frame.close()
    elif db_mode == "店舗":
        conn_fukuoka = sqlite3.connect("framefukuoka.db")
        try:
            df_orders = pd.read_sql("SELECT * FROM orders", conn_fukuoka)
        except Exception as e:
            messagebox.showerror("エラー", f"店舗用DB読込エラー: {e}")
            conn_prod.close()
            conn_fukuoka.close()
            return
        conn_fukuoka.close()
    elif db_mode == "ALL":
        conn_frame = sqlite3.connect("frame.db")
        conn_fukuoka = sqlite3.connect("framefukuoka.db")
        try:
            df_orders_web = pd.read_sql("SELECT * FROM orders", conn_frame)
            df_orders_shop = pd.read_sql("SELECT * FROM orders", conn_fukuoka)
            df_orders = pd.concat([df_orders_web, df_orders_shop], ignore_index=True)
        except Exception as e:
            messagebox.showerror("エラー", f"DB読込エラー: {e}")
            conn_prod.close()
            conn_frame.close()
            conn_fukuoka.close()
            return
        conn_frame.close()
        conn_fukuoka.close()
    else:
        messagebox.showerror("エラー", "データ種別を選択してください。")
        conn_prod.close()
        return

    # --- 以降は今まで通り ---
    # 年月キー追加
    df_orders["年"] = pd.to_numeric(df_orders["年"], errors="coerce")
    df_orders["月"] = pd.to_numeric(df_orders["月"], errors="coerce")
    df_orders = df_orders.dropna(subset=["年", "月"])
    df_orders["年月"] = df_orders["年"].astype(int).astype(str).str[-2:] + "/" + df_orders["月"].astype(int).astype(str).str.zfill(2)
    df_orders["年月キー"] = (df_orders["年"].astype(int) * 100 + df_orders["月"].astype(int)).astype("Int64")
    df_orders = df_orders[(df_orders["年月キー"] >= start_key) & (df_orders["年月キー"] <= end_key)]

    # --- ここから品番範囲フィルタをproduct.dbで実施 ---
    if start_product_code and end_product_code:
        df_products = df_products[
            (df_products["品番CD"].str[:8] >= start_product_code[:8]) &
            (df_products["品番CD"].str[:8] <= end_product_code[:8])
        ]
    # --- ここまで ---

    # 1. 革の種類が"SAMPL"のものは除外
    df_products = df_products[df_products["革の種類"] != "SAMPL"]
    # 金額が空（NaNやNone）のものも除外
    if "金額" in df_products.columns:
        df_products = df_products[df_products["金額"].notna()]

    # frame側の品番10桁列とproduct側の結合キー列を作成
    df_orders["品番10"] = df_orders["品番"].astype(str).str[:10]
    df_products["結合キー"] = df_products["品番CD"].astype(str) + "0" + df_products["カラーNO"].astype(str)

    # 月ごとにワイド形式で集計
    df_orders["点数"] = pd.to_numeric(df_orders["点数"], errors="coerce").fillna(0)
    summary = (
        df_orders.groupby(["品番10", "年月"], as_index=False)["点数"].sum()
    )
    summary_pivot = summary.pivot(index="品番10", columns="年月", values="点数")
    summary_pivot = summary_pivot.fillna(0).astype(int)
    summary_pivot.reset_index(inplace=True)

    # 商品マスタと結合
    merged = pd.merge(
        df_products,
        summary_pivot,
        left_on="結合キー",
        right_on="品番10",
        how="left"
    )

    # 商品マスタに無い品番10桁を抽出
    not_in_master = set(df_orders["品番10"]) - set(df_products["結合キー"])
    if not_in_master:
        # 上2桁が"71"のものだけ抽出
        not_in_master_71 = sorted([code for code in not_in_master if str(code).startswith("71")])
        msg = f"商品マスタに存在しない品番（10桁, 先頭2桁=71）: {not_in_master_71}"
        print(msg)
        # messagebox.showinfo("商品マスタに無い品番(71)", msg)  # 必要ならコメントアウト解除

    # 列順序
    base_cols = ["品番CD", "商品名", "カラーNO", "カラー名", "サイズ数計", "金額", "革の種類"]
    month_cols = [col for col in summary_pivot.columns if col != "品番10"]
    output_cols = base_cols + month_cols
    merged = merged[output_cols]

    # 売上が0の場合は0を入れる
    merged[month_cols] = merged[month_cols].fillna(0).astype(int)

    # === 点数合計列を追加 ===
    merged["点数合計"] = merged[month_cols].sum(axis=1)
    output_cols = base_cols + month_cols + ["点数合計"]
    merged = merged[output_cols]
    # =======================

    # 3. ソート
    leather_order = ["BRI", "VIN BR", "NATUR", "VTC BADALASSI", "BADALASSI", "BR/RUS"]
    merged["__leather_order"] = merged["革の種類"].apply(
        lambda x: leather_order.index(x) if x in leather_order else len(leather_order)
    )
    merged["__品番4"] = merged["品番CD"].str[:4]
    merged = merged.sort_values(
        by=["__品番4", "__leather_order", "革の種類", "品番CD"],
        ascending=[True, True, True, True]
    )

    # 4. ソート後、同じ並びの重複分は指定カラムのみnanにする
    dup_cols = ["品番CD", "商品名", "金額"]
    mask = merged.duplicated(subset=dup_cols, keep='first')
    for col in dup_cols:
        merged.loc[mask, col] = pd.NA

    # ファイル名と保存先
    folder = os.path.join(os.path.expanduser("~"), "Downloads")
    db_mode_str = db_mode  # "店舗"、"WEB"、"ALL" のいずれか
    filename = f"商品別売上集計{start_year_var.get()}年{start_month_var.get()}月-{end_year_var.get()}年{end_month_var.get()}月_{db_mode_str}.xlsx"
    full_path = os.path.join(folder, filename)

    # Excel出力（以降は既存のまま）
    try:
        merged.to_excel(full_path, index=False)

        # openpyxlで色付け
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill

        wb = load_workbook(full_path)
        ws = wb.active

        # 色分け用に「革の種類」情報を一時保存
        leather_types = merged["革の種類"] if "革の種類" in merged.columns else None
        # もし削除済みなら、色分け前にleather_types = merged["革の種類"].copy() で保存しておく

        # === ここで不要な列を削除して再出力 ===
        drop_cols = ["カラーNO", "サイズ数計", "革の種類", "__leather_order", "__品番4"]
        for col in drop_cols:
            if col in merged.columns:
                merged.drop(columns=col, inplace=True)
        merged.to_excel(full_path, index=False)

        # openpyxlで色付け（この時点で「革の種類」列はもう無いので、色分け用データを事前に保存しておく必要あり）
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill

        wb = load_workbook(full_path)
        ws = wb.active

        # ヘッダー行から「品番CD」の列番号を取得
        header = [cell.value for cell in ws[1]]
        col_idx品番CD = header.index("品番CD") + 1  # openpyxlは1始まり

        # 塗りつぶしスタイル（革の種類ごとに色を指定）
        fill_colors = {
            "BRI": "EBF1DE",           # RGB(235,241,222)
            "VIN BR": "CCC0DA",        # RGB(204,192,218)
            "NATUR": "FDE9D9",         # RGB(253,233,217)
            "VTC BADALASSI": "DDD9C4", # RGB(221,217,196)
            "BADALASSI": "FABF8F",     # RGB(250,191,143)
        }

        # 色分け（leather_typesを使う）
        if leather_types is not None:
            for row in range(2, ws.max_row + 1):
                leather_type = leather_types.iloc[row - 2]  # pandasのindexは0始まり
                if leather_type in fill_colors:
                    fill = PatternFill(start_color=fill_colors[leather_type], end_color=fill_colors[leather_type], fill_type="solid")
                    ws.cell(row=row, column=col_idx品番CD).fill = fill

        # === ここからフォント設定追加 ===
        from openpyxl.styles import Font

        # ヘッダー行から対象列のインデックス取得
        header = [cell.value for cell in ws[1]]
        col_idx_map = {name: idx + 1 for idx, name in enumerate(header)}
        # 品番CD・カラー名・金額はフォントサイズ9
        for col_name in ["品番CD", "カラー名", "金額"]:
            if col_name in col_idx_map:
                col_idx = col_idx_map[col_name]
                for row in range(2, ws.max_row + 1):  # 2行目から最終行まで
                    ws.cell(row=row, column=col_idx).font = Font(size=9)
        # 商品名・点数はBOLD
        for col_name in ["商品名", "点数"]:
            if col_name in col_idx_map:
                col_idx = col_idx_map[col_name]
                for row in range(1, ws.max_row + 1):  # ヘッダーも含めてBOLD
                    old_font = ws.cell(row=row, column=col_idx).font
                    ws.cell(row=row, column=col_idx).font = Font(
                        name=old_font.name if old_font else None,
                        size=old_font.size if old_font and old_font.size else None,
                        bold=True
                    )
        # === ここまで追加 ===

        # === ここから罫線追加 ===
        from openpyxl.styles import Border, Side

        thin = Side(border_style="thin", color="000000")
        # まず全セルの罫線をクリア
        #for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
         #   for cell in row:
          #      cell.border = Border()

        # ヘッダー行には下罫線
        #for cell in ws[1]:
            #cell.border = Border(bottom=thin)

        # 「品番CD」列のインデックスを取得
        header = [cell.value for cell in ws[1]]
        try:
            col_idx品番CD = header.index("品番CD") + 1  # openpyxlは1始まり
        except ValueError:
            col_idx品番CD = 1  # デフォルトで1列目

        # 2行目以降、「品番CD」がNaNまたは直前行と異なる場合、その行全体に上罫線
        prev_value = ws.cell(row=1, column=col_idx品番CD).value
        for row in range(2, ws.max_row + 1):
            curr_value = ws.cell(row=row, column=col_idx品番CD).value
            if (prev_value is None and curr_value is not None) or \
               (prev_value is not None and curr_value is not None and prev_value != curr_value):
                # その行全体に上罫線
                for col in range(1, ws.max_column + 1):
                    old_border = ws.cell(row=row, column=col).border
                    ws.cell(row=row, column=col).border = Border(
                        left=thin,
                        right=thin,
                        top=thin,
                        bottom=old_border.bottom,
                    )
            else:
                # それ以外の行も左右罫線は必ず引く
                for col in range(1, ws.max_column + 1):
                    old_border = ws.cell(row=row, column=col).border
                    ws.cell(row=row, column=col).border = Border(
                        left=thin,
                        right=thin,
                        top=old_border.top,
                        bottom=old_border.bottom,
                    )
            prev_value = curr_value

        # 最終行の下罫線を追加
        for col in range(1, ws.max_column + 1):
            old_border = ws.cell(row=ws.max_row, column=col).border
            ws.cell(row=ws.max_row, column=col).border = Border(
                left=old_border.left if old_border.left else thin,
                right=old_border.right if old_border.right else thin,
                top=old_border.top,
                bottom=thin,
            )
        # === ここまで追加 ===

        wb.save(full_path)

        messagebox.showinfo("完了", f"✅ 商品別売上集計を出力しました：\n{full_path}")
    except Exception as e:
        messagebox.showerror("エラー", f"Excel出力エラー: {e}")

    conn_prod.close()

def open_upload_window():
    upload_window = tk.Toplevel()
    upload_window.title("アップロードメニュー")
    upload_window.geometry("400x500")

    tk.Label(upload_window, text="WEB用のExcelファイルを選択してアップロード", font=("Arial", 10)).pack(pady=5)
    tk.Button(upload_window, text="ファイルをアップロード", command=upload_file).pack(pady=10)
    tk.Button(upload_window, text="複数ファイルをアップロード", command=upload_files).pack(pady=10)

    tk.Label(upload_window, text="店舗用Excelファイルを選択してアップロード", font=("Arial", 10)).pack(pady=5)
    tk.Button(upload_window, text="ファイルをアップロード", command=upload_fukuoka_file).pack(pady=10)
    tk.Button(upload_window, text="複数ファイルをアップロード", command=upload_fukuoka_files).pack(pady=10)

    tk.Label(upload_window, text="【商品マスタ管理】", font=("Arial", 10, "bold")).pack(pady=10)
    tk.Button(upload_window, text="商品マスタExcel取込", command=upload_product_excel).pack(pady=5)
    tk.Button(upload_window, text="商品マスタ金額CSV取込", command=upload_price_csv).pack(pady=5)

    # 閉じるボタン
    tk.Button(upload_window, text="閉じる", command=upload_window.destroy).pack(pady=20)

# 統合GUI
window = tk.Tk()
window.title("売上データ処理システム")
window.geometry("400x500")

tk.Button(window, text="アップロード", command=open_upload_window).pack(pady=20)

tk.Label(window, text="年月指定してデータをExcelで出力", font=("Arial", 10)).pack(pady=10)

from datetime import datetime
current_year = datetime.now().year
current_month = datetime.now().month
years = list(range(2014, current_year + 1))
months = list(range(1, 13))
start_year_var = tk.StringVar(value=str(current_year))
start_month_var = tk.StringVar(value=str(current_month - 1))
end_year_var = tk.StringVar(value=str(current_year))
end_month_var = tk.StringVar(value=str(current_month - 1))

tk.Label(window, text="開始 年月").pack()
frame1 = tk.Frame(window)
frame1.pack()
ttk.Combobox(frame1, textvariable=start_year_var, values=years, width=6).pack(side="left")
ttk.Combobox(frame1, textvariable=start_month_var, values=months, width=4).pack(side="left")

tk.Label(window, text="終了 年月").pack()
frame2 = tk.Frame(window)
frame2.pack()
ttk.Combobox(frame2, textvariable=end_year_var, values=years, width=6).pack(side="left")
ttk.Combobox(frame2, textvariable=end_month_var, values=months, width=4).pack(side="left")

# "品番" 範囲指定用テキストボックス
tk.Label(window, text="③ 品番を指定（範囲指定: 12桁）").pack(pady=10)
frame3 = tk.Frame(window)
frame3.pack()
tk.Label(frame3, text="開始品番:").pack(side="left")
start_product_code_var = tk.StringVar()
tk.Entry(frame3, textvariable=start_product_code_var, width=15).pack(side="left")
tk.Label(frame3, text="終了品番:").pack(side="left")
end_product_code_var = tk.StringVar()
tk.Entry(frame3, textvariable=end_product_code_var, width=15).pack(side="left")

tk.Button(window, text="Excelファイルに出力", command=lambda: export_data(start_product_code_var.get(), end_product_code_var.get())).pack(pady=20)

# === ラジオボタン追加 ===
db_select_var = tk.StringVar(value="ALL")  # デフォルトはWEB
radio_frame = tk.Frame(window)
radio_frame.pack(pady=5)
tk.Radiobutton(radio_frame, text="店舗", variable=db_select_var, value="店舗").pack(side="left", padx=10)
tk.Radiobutton(radio_frame, text="WEB", variable=db_select_var, value="WEB").pack(side="left", padx=10)
tk.Radiobutton(radio_frame, text="ALL", variable=db_select_var, value="ALL").pack(side="left", padx=10)

tk.Button(window, text="商品別売上集計ダウンロード", command=download_sales_summary).pack(pady=10)

# 閉じるボタン
tk.Button(window, text="閉じる", command=window.destroy).pack(pady=20)

window.mainloop()